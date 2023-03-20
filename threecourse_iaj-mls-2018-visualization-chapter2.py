# osモジュールのimport
import os

# numpyモジュールをnpという名前で読み込む
import numpy as np

# sklearn.ensembleモジュールにあるRandomForestClassifierを読み込む
from sklearn.ensemble import RandomForestClassifier
import os
print(os.getcwd()) # カレントディレクトリの表示
print(os.path.join("dir1", "subdir1")) # パスの結合
import itertools
for plan, (n, m) in itertools.product(["A", "B"], [(10, 10), (10, 60), (60, 60)]):   
    print(plan, n, m)
# ファイルを読込モードでエンコードUTF-8によって開く
f = open("../input/mlsinput/ch2-1.txt", "r", encoding="UTF-8")
# ファイルの中身を変数に格納する
lines = f.readlines()
# 付加されている改行文字を除去する
lines = [line.rstrip() for line in lines]
# ファイルを閉じる
f.close()

# 各行ごとに処理する
for line in lines:
    print(line)

# 行を空白などで分割して使うこともできます。 
words = lines[0].split(' ')
print(words)
# 行ごとのリストとして書き込みたいデータを準備する
lines = ["1: これは１行目です",  "2: This is 2nd line."]
# ファイルを書込モードで開く
f = open("ch2-out1.txt", "w", encoding="utf-8")
# 改行を間に入れて各行を結合した文字列を作り、それを書き込む
f.write("\n".join(lines) + "\n")
# ファイルを閉じる
f.close()
import openpyxl

# ワークブックを開く
wb = openpyxl.load_workbook(filename="../input/mlsinput/ch2-1.xlsx", read_only=True)

# ワークシートの指定
ws = wb['first']

# 行ごとに処理
for row in ws.rows:
    cell_values = []
    
    # 各行のセルごとに値を読み込む
    for cell in row:        
        cell_values.append(cell.value)
        
    # カンマで繋げて出力    
    print(", ".join(cell_values))    

# ワークブックを閉じる    
wb.close()
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# ワークブックを作成
wb = openpyxl.Workbook()

# ワークシートの指定
ws = wb.active

# ワークシート名の変更
ws.title = "first"

# セルごとに処理
for r in range(1, 6):
    for c in range(1, 6):
        # セルを作成
        cell = ws.cell(column=c, row=r, value="{}-{}".format(r, c))

        # セルのフォント・揃え・色を変更
        cell.font = Font(name='メイリオ', bold=False, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if r == 1:
            cell.fill = PatternFill("solid", fgColor="D8E4BC")            

# ワークブックを保存して閉じる
wb.save("ch2-out1.xlsx")
wb.close()
# モンスターの元となるクラス
class Monster:
    
    def __init__(self):    
        self.name = ""
        self.hp = 0
        self.atk = 0

    def attack(self):
        print("{}は{}ダメージを与える".format(self.name, self.atk))

# モンスタークラスを継承したスライムクラス
class Slime(Monster):
    
    def __init__(self):    
        self.name = "スライム"
        self.hp = 5
        self.atk = 1

# モンスタークラスを継承したドラゴンクラス
class Dragon(Monster):
    
    def __init__(self):
        self.name = "ドラゴン"
        self.hp = 50
        self.atk = 10
        
    def fire_breath(self):
        print("{}は全体に{}ダメージを与える".format(self.name, self.atk))
    
monster1 = Slime()
monster1.attack()

monster2 = Dragon()
monster2.attack()
monster2.fire_breath()
# レートの元となるクラス
class Rate:
    def __init__(self, x):
        self.x = x
    def premium(self):
        pass
    def reserve(self):
        pass
    
# 養老保険のレート
class EndowRate(Rate):
    def premium(self):
        return 0.1 + self.x * 0.002
    def reserve(self):
        return [0.0, 0.2, 0.4, 0.6, 0.8, 1.0]

# 保険料を出力する関数   
def calculate_premium(rate):
    print(rate.premium())
    
calculate_premium(EndowRate(20))
from datetime import datetime

# ユーティリティ的な関数をまとめておくクラス
class Utility:
    
    # 見た目を整備して現在時刻を出力する
    @classmethod
    def pretty_print_now(cls):
        print(datetime.now().strftime("[%Y-%m-%d %H:%M:%S]"))
        
Utility.pretty_print_now()
# YOUR CODE GOES HERE
# YOUR CODE GOES HERE