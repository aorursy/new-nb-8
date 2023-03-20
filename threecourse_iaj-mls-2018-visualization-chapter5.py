# 各ライブラリのimport
import numpy as np
import pandas as pd
import openpyxl
# ヘッダがある場合
df = pd.read_excel("../input/mlsinput/ch5-1.xlsx", "inforce")
df
# ヘッダが無い場合
df = pd.read_excel("../input/mlsinput/ch5-1.xlsx", "inforce2", header=None, names=["ID", "plan", "x", "n", "m", "S"])
df
# テーブルの外側にコメントを記述したセルがある場合
df = pd.read_excel("../input/mlsinput/ch5-1.xlsx", "inforce3", header=None)
df
# https://openpyxl.readthedocs.io/en/stable/tutorial.html

from openpyxl import load_workbook

# ワークシートの取得
# 算式でなく値を取得するためには、data_only=Trueとする必要がある
wb = load_workbook(filename = "../input/mlsinput/ch5-1.xlsx", data_only=True)

# シート名一覧の取得
print ("シート", wb.sheetnames)

# シートの取得
ws = wb['inforce']

# セルの値の取得（範囲の名前）
value = ws['A4'].value
print("value A4 -", value)

# セルの値の取得（行数・列数は1始まり）
value = ws.cell(row=4, column=1).value
print("value A4 -", value)

# 使用している行数・列数の取得
print("rows: {}, columns: {}".format(ws.max_row, ws.max_column))

# シートの全体の値を取得
data = list(ws.values)
cols = data[0]
values = data[1:]
print(cols)
print(values)

# pandasのデータフレームへの変換
df = pd.DataFrame(values, columns=cols)
df
ary = np.arange(24).reshape(4, 6)
df = pd.DataFrame(ary)
df.to_excel("out5-1.xlsx")
df.to_csv("out5-1.txt", index=False, sep="\t")
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelUtil(object):

    @classmethod
    def add_sheet(cls, _wb, _df, sheet_name, index_cols_count, 
                  value_fmt="#,##0", value_centered=False, index_width=7, value_width=11):
        """
        ワークブックにデータフレームのデータから作成したワークシートを追加する
        書式も設定する
        """
        ws = _wb.create_sheet(sheet_name)
        for r in dataframe_to_rows(_df, index=False, header=True):
            ws.append(r)
        cls.set_style_worksheet(ws, index_cols_count, value_fmt, value_centered, index_width, value_width)
            
        # 一応workbookとworksheetを返り値として返しておく
        return wb, ws

    @classmethod
    def set_style_worksheet(cls, ws, icols, value_fmt, value_centered, index_width, value_width):
        """
        ワークシートに書式を設定する
        """
        cs = len(list(ws.columns))
        rs = len(list(ws.rows))

        font = Font(name='メイリオ', bold=False, size=11)
        bd = Side(style='thin', color="000000")
        border = Border(left=bd, top=bd, right=bd, bottom=bd)
        alignment = Alignment(horizontal='center', vertical='center')
        fill = PatternFill("solid", fgColor="D8E4BC")
        if value_centered:
            value_alignment = Alignment(horizontal='center') 
        else:
            value_alignment = None
        
        # indexの書式
        if icols > 0:
            cls.set_style_range(ws, 1, rs, 1, icols, 
                                font=font, border=border, alignment=alignment, fill=fill, format="General")
        
        # headerの書式
        cls.set_style_range(ws, 1, 1, 1, cs, 
                            font=font, border=border, alignment=alignment, fill=fill, format="General")
        
        # valueの書式
        cls.set_style_range(ws, 2, rs, icols+1, cs,
                            font=font, border=border, alignment=value_alignment, fill=None, format=value_fmt)
        
        # 列の幅
        cls.set_column_width(ws, 1, icols, index_width)
        cls.set_column_width(ws, icols+1, cs, value_width)
    
    @classmethod
    def set_style_range(cls, ws, r_start, r_last, c_start, c_last, font=None, border=None, alignment=None, fill=None, format=None):
        """
        範囲に書式を設定する
        """

        cs = len(list(ws.columns))
        rs = len(list(ws.rows))
        for r in range(rs):
            for c in range(cs):
                rr = r + 1
                cc = c + 1
                if r_start <= rr <= r_last and c_start <= cc <= c_last:
                    cell = ws.cell(column=c + 1, row=r + 1)
                    if font is not None: cell.font = font
                    if border is not None: cell.border = border
                    if alignment is not None: cell.alignment = alignment
                    if fill is not None: cell.fill = fill
                    if format is not None: cell.number_format = format
        return ws

    @classmethod
    def set_column_width(cls, ws, c_start, c_last, width):
        """
        列の幅を指定する
        """
        for c in range(c_start, c_last+1):
            ws.column_dimensions[get_column_letter(c)].width = width
    
df_inforce = pd.read_csv("../input/mlsinput/insurance_inforce.csv")
df_summary = pd.read_csv("../input/mlsinput/insurance_sensitivity.csv") 
wb = openpyxl.Workbook()
wb, ws = ExcelUtil.add_sheet(wb, df_inforce, "inforce", 1, "General", value_centered=True, index_width=10)
wb, ws = ExcelUtil.add_sheet(wb, df_summary, "summary", 3, "0.00%", index_width=10)
wb.remove_sheet(wb.get_sheet_by_name(wb.sheetnames[0]))
wb.save("out5-2.xlsx")
# （Excelが入っていない環境では動きません）
import xlwings as xw
wb = xw.Book(r'..\input\mlsinput\ch5-2.xlsm')
app = xw.apps.active    
macro = wb.macro('Run')
result = macro() # VBAの返り値を取得できる
print(result)
wb.save()
wb.close()
app.kill() 